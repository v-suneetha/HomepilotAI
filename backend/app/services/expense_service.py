from __future__ import annotations

import csv
import io
from typing import List

from openpyxl import load_workbook

from app.repositories.expense_repository import ExpenseRepository


class ExpenseService:
    def __init__(self, repository: ExpenseRepository | None = None):
        self.repository = repository or ExpenseRepository()

    def import_csv(self, file_bytes: bytes) -> dict:
        text_data = file_bytes.decode("utf-8")
        reader = csv.DictReader(io.StringIO(text_data))
        imported_count = 0
        for row in reader:
            amount = float(row.get("amount", 0))
            description = row.get("description", "") or ""
            date = row.get("date", "") or ""
            category_name = row.get("category") or None
            self.repository.add_transaction(description=description, amount=amount, date=date, category_name=category_name)
            imported_count += 1
        return {"imported_count": imported_count, "categories_created": 1 if imported_count else 0}

    def import_excel(self, file_bytes: bytes) -> dict:
        workbook = load_workbook(filename=io.BytesIO(file_bytes), data_only=True, read_only=True)
        sheet = workbook.active
        rows = sheet.iter_rows(values_only=True)
        header = next(rows, None)
        imported_count = 0
        if header is None:
            return {"imported_count": 0, "categories_created": 0}

        for row in rows:
            row_data = dict(zip(header, row))
            amount = float(row_data.get("amount", 0) or 0)
            description = str(row_data.get("description", "") or "")
            date = str(row_data.get("date", "") or "")
            category_name = str(row_data.get("category", "") or "") or None
            self.repository.add_transaction(description=description, amount=amount, date=date, category_name=category_name)
            imported_count += 1
        return {"imported_count": imported_count, "categories_created": 1 if imported_count else 0}

    def summary(self) -> dict:
        return self.repository.get_summary()

    def set_budget(self, category_name: str, monthly_limit: float, month: str | None = None) -> dict:
        budget = self.repository.set_budget(category_name, monthly_limit, month)
        return {
            "category": budget.category_name,
            "limit": round(budget.monthly_limit, 2),
            "month": budget.month,
        }
